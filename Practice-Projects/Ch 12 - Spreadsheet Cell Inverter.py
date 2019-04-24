#! python3
# cell inverter - transposes cells on an excel file
"""
Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
This should be done for all cells in the spreadsheet.
"""
import openpyxl
import os

# Open excel file as well as a new empty one to write to.
# Create two empty lists. We will start each row from the first workbook into a list
# then output both lists into the new excel file
input_wb = openpyxl.load_workbook('produce_sales.xlsx')
input_sheet = input_wb.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
col_1 = []
col_2 = []

# This nested for loop will add each column to the empty lists. In his example, there are only two columns.
# You'll have an if statement that places the value from one column into one list, then the second column
# in the next list
for row in range(1, input_sheet.max_row + 1):
    for col in range(1, input_sheet.max_column + 1):
        if col is 1:
            col_val_1 = input_sheet.cell(row=row, column=col).value
            col_1.append(col_val_1)
        else:
            col_val_2 = input_sheet.cell(row=row, column=col).value
            col_2.append(col_val_2)


# For transposing, we need a different kind of loop. We are looping through each item in the list and we create
# our own iterator for the column. Set each value to have row=1 and add the column number after each value
# is copied to the new sheet. Repeat for the second list, but row=2.
col_count = 1
col_count_2 = 1
for item in col_1:
    output_sheet.cell(row=1, column=col_count).value = item
    col_count += 1
for item in col_2:
    output_sheet.cell(row=2, column=col_count_2).value = item
    col_count_2 += 1
os.chdir(r'c:\users\tomal\desktop')
output_wb.save('post_cell_invert.xlsx')
