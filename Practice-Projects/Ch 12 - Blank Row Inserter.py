#! python3
# blank_row_inserter - inserts blank rows into excel file
"""
Create a program blankRowInserter.py that takes two integers and a filename string as command line
arguments. Let’s call the first integer N (row_start) and the second integer M (row_length). Starting at row N, the program
should insert M blank rows into the spreadsheet.
"""

import sys
import os
import openpyxl


row_start = int(sys.argv[1])
row_length = int(sys.argv[2])
filename = sys.argv[3]
# Open the excel file you want to edit as well as a new one. Remember, we don't EDIT the original,
# but we write our info to a new spreadsheet.
print('Opening workbook...')
os.chdir(r'C:\Users\tomal\Desktop\Automate The Boring Stuff Files\automate_online-materials')
wb = openpyxl.load_workbook(filename)
sheet = wb.active
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

print('Adding blank rows...')
# Once again, multiple for loops!
# Start from row/column since excel doesn't have row/column 0
# Add 1 to the max row and column since we are adding 1 from 0
# Copy the rows like normal until you get the row_start, then start adding the row_length
# in each loops. This will "skip" the rows in the command line argument.
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        if row < row_start:
            new_sheet.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
        else:
            new_sheet.cell(row=row+row_length, column=column).value = sheet.cell(row=row, column=column).value


os.chdir(r'c:\users\tomal\desktop')
new_wb.save('blank_row.xlsx')
print('Done.')
