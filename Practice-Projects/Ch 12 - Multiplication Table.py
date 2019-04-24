#! python3
# multiplication_table.py - creates a multiplication table in excel with N rows
"""
Create a program multiplicationTable.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet.

Note with this module! There have been updates to openpyxl since Automate The Boring Stuff
was written and it has caused a lot of syntax to change. I'll try to comment out each part,
but I apologize if I miss any
"""

import sys
import openpyxl
import os
# Notice, we have to use openpyxyl.utils instead of openpyxl.cell now
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

os.chdir(r'c:\users\tomal\desktop')

# sys.argv[1] will be the first argument entered into the command line
# Remember, sys.argv[0] is the first argument and it's a string
# containing the filename! Don't forget to convert the argument into an int
n_table = int(sys.argv[1])

# Open a new workbook and make your sheet active.
wb = openpyxl.Workbook()
sheet = wb.active

# This part can get tricky. It may take a lot of guessing and checking to ensure you do it
# correctly with the number of rows and columns. You want to start from row 2 since cell A1 is empty
# This first for loop will write out the rows and columns up to the number that was entered in the
# command line. When trying to add a value to a cell you need a letter and number, but in string format.
# So pass str() to your r iterator.
# The value will be r - 1, since we are starting on row 2.
for r in range(2, n_table + 2):
    sheet['A' + str(r)] = r - 1
    col_let = get_column_letter(r)
    sheet[col_let + str(1)] = r - 1
    # Set font to bold
    sheet['A' + str(r)].font = Font(bold=True)
    sheet[col_let + str(1)].font = Font(bold=True)

# This is the loop that will create the values inside the box created from above. You will need two loops
# One for each row, one for each column
# First, iterate over each row in one column, then move on the the next column, and so on...
# Multiply the column * row (subract 1 from both) since we are starting from row 2
for col in range(2, n_table + 2):
    for row in range(2, n_table + 2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col - 1) * (row - 1)

wb.save('Multiplication Table.xlsx')
