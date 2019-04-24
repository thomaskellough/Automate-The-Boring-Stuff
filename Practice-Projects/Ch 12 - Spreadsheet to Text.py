#! python3
# spreadsheet to text.py - will import an excel sheet and export it as multiple txt documents
"""
Write a program that performs the tasks of the previous program in reverse order:
The program should open a spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.

Note: For this program I stored a list inside another list. However, because of the loops, some of
the inner lists were left with None values. After much Google searching, I discovered a way to remove
these items. Example below:

spam = [1, 2, 3, 4, None, None, 5, 6, None]
print(spam)
spam = [x for x in spam if x is not None]
print(spam)

In this example, x is just a variable and can be named anything. You are replacing each item in
the list with itself, but ony if the value is not None. This will create a new list without None values
"""

import openpyxl


wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb.active

# Create two empty lists. One for the text files, and one to hold the list of text files
text = []
full_list = []
# Read each cell and append each string inside a list, then append that list to another list.
# In the end, you'll end up with one list that contains multiple other lists. The inner lists are
# strings of text read from the excel spreadsheet. Remeber, after appending each inner list to the
# outer list, you need to set it to empty again for the new column
for col in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        text.append(sheet.cell(row=row, column=col).value)
        text = [x for x in text if x is not None]
    full_list.append(text)
    text = []


# This is, once again, another double for loop. (Tired of them yet?)
# The first for loop iterates over each list inside full_list.
# The second for loop iterates over each string inside each inner list.
# Open a new text file in 'w' mode for each inner list, then write each string.
# of the inner list inside the text file. Remember to convert (item) to string.
# Separate them with a new line. Then add 1 to the column
# and close the text file AFTER each inner list is complete.
col = 1
for list in full_list:
    text_file = open('spread_to_text' + str(col) + '.txt', 'w')
    for item in list:
        text_file.write(str(item) + '\n')
    col += 1
    text_file.close()
